import pytest

def test_example():
    assert 1 + 1 == 2

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import datetime

class OrangeHRMLoginPage:
    def __init__(self, driver):
        self.driver = driver
        self.username_input = (By.ID, "txtUsername")
        self.password_input = (By.ID, "txtPassword")
        self.login_button = (By.ID, "btnLogin")

    def login(self, username, password):
        self.driver.get("https://opensource-demo.orangehrmlive.com/")
        WebDriverWait(self.driver, 10).until(EC.visibility_of_element_located(self.username_input)).send_keys(username)
        self.driver.find_element(*self.password_input).send_keys(password)
        self.driver.find_element(*self.login_button).click()
        return EC.visibility_of_element_located((By.ID, "welcome"))(self.driver)

def run_tests():
    excel_file = "test_data.xlsx"
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    driver = webdriver.Chrome()
    login_page = OrangeHRMLoginPage(driver)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        test_id, username, password, _, _, tester_name, _ = row
        result_col = sheet.max_column

        try:
            if login_page.login(username, password):
                sheet.cell(row=test_id, column=result_col).value = "Pass"
            else:
                sheet.cell(row=test_id, column=result_col).value = "Fail"
        except Exception as e:
            print(f"Test ID {test_id} failed with error: {str(e)}")
            sheet.cell(row=test_id, column=result_col).value = "Fail"

    wb.save(excel_file)
    driver.quit()

if __name__ == "__main__":
    test_example()